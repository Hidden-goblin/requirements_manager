# -*- coding: utf-8 -*-
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from openpyxl import load_workbook
from datetime import datetime

from openpyxl.utils import get_column_letter, column_index_from_string


def import_testinsight(parent = None):
    try:
        file_to_import = QFileDialog.getOpenFileName(parent, 'Open File', "", "Excel Spreadsheet (*.xlsx)")[0]
        starting_time = "{0:%Y}{0:%m}{0:%d}{0:%H}{0:%M}{0:%S}".format(datetime.today())
        mapping = {'Functionality Name': -1,
                   "Functionality Channel": 3,
                   "Use case name": 4,
                  "Test step": 7,
                  "Expected Result": 8,
                  "Pre Requisite": 10}
        workbook = load_workbook(file_to_import)
        worksheet = workbook.get_sheet_by_name('Test Case')
        if worksheet.cell(row = 4, column = mapping['Functionality Channel']).value != 'Requirement Name' or worksheet.cell(row = 4, column = mapping['Use case name']).value != 'Test Title' or worksheet.cell(row = 4, column = mapping['Test step']).value != 'Test Step' or worksheet.cell(row = 4, column = mapping['Expected Result']).value != 'Expected Result' or worksheet.cell(row = 4, column = mapping['Pre Requisite']).value != 'Pre Requisite':
            raise KeyError("Spreadsheet haven't the expected layout.\n It must be on row 4 (header/column):\n "
                       "Requirement Name/C, Test Title/D, Test Step/G, Pre Requisite/J, and Expected Result/H ")

        funct_count = 0
        for row in range(5, worksheet.max_row + 1):
            if worksheet.cell(row = row, column = mapping['Functionality Channel']).value:
                funct_count += 1
                current_functionality_id = starting_time + str(row)
                dictionary = {"ID": current_functionality_id,
                              "Name": "Imported form TestInsight",
                              "Description": "Imported from TestInsight",
                              "Channel": worksheet.cell(row = row, column = mapping['Functionality Channel']).value,
                              "Criticality": 1,
                              "Usage": 1,
                              "Screen": ""}
                parent.repository.add_functionality(functionality = dictionary)
                del dictionary
            #print("Create channel '{}' with ID '{}' and description 'Import form TestInsight'".format(worksheet.cell(row = row, column = mapping['Functionality Channel']).value, current_functionality_id))
            if worksheet.cell(row = row, column = mapping['Use case name']).value:
                current_use_case_id = starting_time + str(row)
                usecase = {"ID": current_use_case_id,
                           "Name": worksheet.cell(row = row, column = mapping['Use case name']).value,
                           "Short Description": "Imported from TestInsight",
                           "Pre Requisite": worksheet.cell(row = row, column = mapping['Pre Requisite']).value,
                           "Steps": worksheet.cell(row = row, column = mapping['Test step']).value,
                            "Post Condition": worksheet.cell(row = row, column = mapping['Expected Result']).value,
                           "Exceptional Behaviour":"",
                           "Require Use Case":"",
                           "Test Type": 0,
                           "Critical Path": False,
                           "Average Run Time": '',
                           "Automated": False,
                           "Automatable": False,
                           "Release": '',
                           "Functionality link": current_functionality_id,
                           "Automaton link": ''}
                parent.repository.add_use_case(use_case = usecase)
                del usecase
            else:
                print("Add values to usecase id '{}'".format(current_use_case_id))

            parent.update_list()
            parent.repository_update_signal.emit(True)
    except Exception as exception:
        print(repr(exception))
